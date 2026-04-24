#!/usr/bin/env python3
"""Cross-check normalized CIK text data against official regional XLSX workbooks."""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


VOTE_SHEET_SUFFIX = "-Гласове"
SECTION_SHEET_SUFFIX = "-Секции"
PROTOCOL_POINTS_SHEET_SUFFIX = "-Точки в протоколи"

PROTOCOL_FIELDS = [
    "received_paper_ballots",
    "registered_voters_initial",
    "voters_added_election_day",
    "voters_signed",
    "unused_paper_ballots",
    "destroyed_paper_ballots",
    "paper_ballots_found",
    "invalid_paper_ballots",
    "paper_none_votes",
    "valid_paper_candidate_votes",
    "machine_ballots_found",
    "machine_none_votes",
    "valid_machine_candidate_votes",
]


def read_dicts(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def to_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, float):
        return int(value)
    return int(str(value).strip())


def normalize_section_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return f"{value:09d}"
    if isinstance(value, float):
        return f"{int(value):09d}"
    return str(value).strip()


def find_sheet(sheetnames: list[str], suffix: str) -> str:
    matches = [name for name in sheetnames if name.endswith(suffix)]
    if len(matches) != 1:
        raise ValueError(f"Expected exactly one sheet ending {suffix!r}, found {matches}")
    return matches[0]


def load_text_sections(processed_dir: Path) -> dict[str, dict[str, Any]]:
    sections = {}
    for row in read_dicts(processed_dir / "sections.csv"):
        section_id = row["section_id"]
        sections[section_id] = {
            "section_id": section_id,
            "region_id": row["region_id"],
            "is_mobile": to_int(row["is_mobile"]),
            "is_ship": to_int(row["is_ship"]),
            "has_machine": int(to_int(row["machines_count"]) > 0),
        }
    return sections


def load_text_votes(processed_dir: Path) -> dict[tuple[str, int], dict[str, Any]]:
    totals: dict[tuple[str, int], dict[str, Any]] = defaultdict(
        lambda: {"valid_votes": 0, "paper_votes": 0, "machine_votes": 0, "row_count": 0, "party_names": set()}
    )
    for row in read_dicts(processed_dir / "votes_long.csv"):
        key = (row["section_id"][:2], to_int(row["party_id"]))
        totals[key]["valid_votes"] += to_int(row["valid_votes"])
        totals[key]["paper_votes"] += to_int(row["paper_votes"])
        totals[key]["machine_votes"] += to_int(row["machine_votes"])
        totals[key]["row_count"] += 1
        totals[key]["party_names"].add(row["party_name"])
    return totals


def load_text_protocols(processed_dir: Path) -> dict[str, dict[str, int]]:
    protocols = {}
    for row in read_dicts(processed_dir / "protocols.csv"):
        protocols[row["section_id"]] = {field: to_int(row[field]) for field in PROTOCOL_FIELDS}
    return protocols


def parse_region_id_from_filename(filename: str) -> str:
    match = re.fullmatch(r"ns(\d{2})\.xlsx", Path(filename).name)
    if not match:
        raise ValueError(f"Unexpected spreadsheet filename: {filename}")
    return match.group(1)


def parse_sections_sheet(ws: Any, region_id: str) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_number == 1:
            continue
        section_id = normalize_section_id(row[0])
        if not section_id:
            continue
        rows[section_id] = {
            "section_id": section_id,
            "region_id": region_id,
            "is_mobile": to_int(row[6]),
            "has_machine": to_int(row[7]),
            "is_ship": to_int(row[10]),
        }
    return rows


def parse_votes_sheet(ws: Any, region_id: str) -> dict[tuple[str, int], dict[str, Any]]:
    totals: dict[tuple[str, int], dict[str, Any]] = defaultdict(
        lambda: {"valid_votes": 0, "paper_votes": 0, "machine_votes": 0, "row_count": 0, "party_names": set()}
    )
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_number == 1:
            continue
        section_id = normalize_section_id(row[0])
        if not section_id:
            continue
        ballot_type = str(row[5]).strip()
        party_id = to_int(row[6])
        party_name = str(row[7]).strip()
        votes = to_int(row[8])
        key = (region_id, party_id)

        if ballot_type == "Хартия":
            totals[key]["paper_votes"] += votes
        elif ballot_type == "Машина":
            totals[key]["machine_votes"] += votes
        else:
            raise ValueError(f"Unknown ballot type {ballot_type!r} in {region_id} row {row_number}")

        totals[key]["valid_votes"] += votes
        totals[key]["row_count"] += 1
        totals[key]["party_names"].add(party_name)
    return totals


def parse_protocol_points_sheet(ws: Any) -> dict[str, dict[str, int]]:
    protocols: dict[str, dict[str, int]] = defaultdict(lambda: {field: 0 for field in PROTOCOL_FIELDS})
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_number == 1:
            continue
        section_id = normalize_section_id(row[0])
        if not section_id:
            continue

        received_paper_ballots = row[6]
        if received_paper_ballots not in (None, ""):
            protocols[section_id].update(
                {
                    "received_paper_ballots": to_int(row[6]),
                    "registered_voters_initial": to_int(row[7]),
                    "voters_added_election_day": to_int(row[8]),
                    "voters_signed": to_int(row[9]),
                    "unused_paper_ballots": to_int(row[10]),
                    "destroyed_paper_ballots": to_int(row[11]),
                    "paper_ballots_found": to_int(row[12]),
                    "invalid_paper_ballots": to_int(row[13]),
                    "paper_none_votes": to_int(row[14]),
                    "valid_paper_candidate_votes": to_int(row[15]),
                }
            )
        else:
            protocols[section_id].update(
                {
                    "machine_ballots_found": to_int(row[12]),
                    "machine_none_votes": to_int(row[14]),
                    "valid_machine_candidate_votes": to_int(row[15]),
                }
            )
    return dict(protocols)


def add_totals(
    destination: dict[tuple[str, int], dict[str, Any]],
    source: dict[tuple[str, int], dict[str, Any]],
) -> None:
    for key, values in source.items():
        for column, value in values.items():
            if column == "party_names":
                destination[key][column].update(value)
            else:
                destination[key][column] += value


def crosscheck(spreadsheet_zip: Path, processed_dir: Path, output_dir: Path) -> None:
    text_sections = load_text_sections(processed_dir)
    text_votes = load_text_votes(processed_dir)
    text_protocols = load_text_protocols(processed_dir)

    xlsx_sections: dict[str, dict[str, Any]] = {}
    xlsx_votes: dict[tuple[str, int], dict[str, Any]] = defaultdict(
        lambda: {"valid_votes": 0, "paper_votes": 0, "machine_votes": 0, "row_count": 0, "party_names": set()}
    )
    xlsx_protocols: dict[str, dict[str, int]] = {}
    workbook_summaries: list[dict[str, Any]] = []

    with zipfile.ZipFile(spreadsheet_zip) as zf:
        workbook_names = sorted(name for name in zf.namelist() if name.endswith(".xlsx"))
        for workbook_name in workbook_names:
            region_id = parse_region_id_from_filename(workbook_name)
            workbook = load_workbook(io.BytesIO(zf.read(workbook_name)), read_only=True, data_only=True)

            sections_sheet = workbook[find_sheet(workbook.sheetnames, SECTION_SHEET_SUFFIX)]
            votes_sheet = workbook[find_sheet(workbook.sheetnames, VOTE_SHEET_SUFFIX)]
            protocol_sheet = workbook[find_sheet(workbook.sheetnames, PROTOCOL_POINTS_SHEET_SUFFIX)]

            region_sections = parse_sections_sheet(sections_sheet, region_id)
            region_votes = parse_votes_sheet(votes_sheet, region_id)
            region_protocols = parse_protocol_points_sheet(protocol_sheet)

            xlsx_sections.update(region_sections)
            add_totals(xlsx_votes, region_votes)
            xlsx_protocols.update(region_protocols)

            workbook_summaries.append(
                {
                    "workbook": workbook_name,
                    "region_id": region_id,
                    "section_rows": len(region_sections),
                    "vote_party_rows": sum(values["row_count"] for values in region_votes.values()),
                    "protocol_sections": len(region_protocols),
                }
            )

    section_mismatches: list[dict[str, Any]] = []
    text_section_ids = set(text_sections)
    xlsx_section_ids = set(xlsx_sections)
    for section_id in sorted(text_section_ids - xlsx_section_ids):
        section_mismatches.append({"section_id": section_id, "issue_type": "missing_from_spreadsheet", "field": "", "text_value": "", "spreadsheet_value": ""})
    for section_id in sorted(xlsx_section_ids - text_section_ids):
        section_mismatches.append({"section_id": section_id, "issue_type": "missing_from_text_tables", "field": "", "text_value": "", "spreadsheet_value": ""})
    for section_id in sorted(text_section_ids & xlsx_section_ids):
        for field in ("region_id", "is_mobile", "is_ship", "has_machine"):
            if text_sections[section_id][field] != xlsx_sections[section_id][field]:
                section_mismatches.append(
                    {
                        "section_id": section_id,
                        "issue_type": "section_field_mismatch",
                        "field": field,
                        "text_value": text_sections[section_id][field],
                        "spreadsheet_value": xlsx_sections[section_id][field],
                    }
                )

    vote_mismatches: list[dict[str, Any]] = []
    vote_keys = set(text_votes) | set(xlsx_votes)
    vote_crosscheck_rows = []
    vote_name_mismatches: list[dict[str, Any]] = []
    for region_id, party_id in sorted(vote_keys):
        text_values = text_votes.get((region_id, party_id), {"valid_votes": 0, "paper_votes": 0, "machine_votes": 0, "row_count": 0, "party_names": set()})
        xlsx_values = xlsx_votes.get((region_id, party_id), {"valid_votes": 0, "paper_votes": 0, "machine_votes": 0, "row_count": 0, "party_names": set()})
        text_party_names = " | ".join(sorted(text_values["party_names"]))
        spreadsheet_party_names = " | ".join(sorted(xlsx_values["party_names"]))
        row = {
            "region_id": region_id,
            "party_id": party_id,
            "text_party_names": text_party_names,
            "spreadsheet_party_names": spreadsheet_party_names,
            "text_valid_votes": text_values["valid_votes"],
            "spreadsheet_valid_votes": xlsx_values["valid_votes"],
            "diff_valid_votes": text_values["valid_votes"] - xlsx_values["valid_votes"],
            "text_paper_votes": text_values["paper_votes"],
            "spreadsheet_paper_votes": xlsx_values["paper_votes"],
            "diff_paper_votes": text_values["paper_votes"] - xlsx_values["paper_votes"],
            "text_machine_votes": text_values["machine_votes"],
            "spreadsheet_machine_votes": xlsx_values["machine_votes"],
            "diff_machine_votes": text_values["machine_votes"] - xlsx_values["machine_votes"],
            "text_row_count": text_values["row_count"],
            "spreadsheet_row_count": xlsx_values["row_count"],
        }
        vote_crosscheck_rows.append(row)
        if row["diff_valid_votes"] or row["diff_paper_votes"] or row["diff_machine_votes"]:
            vote_mismatches.append(row)
        if text_party_names != spreadsheet_party_names:
            vote_name_mismatches.append(
                {
                    "region_id": region_id,
                    "party_id": party_id,
                    "text_party_names": text_party_names,
                    "spreadsheet_party_names": spreadsheet_party_names,
                }
            )

    protocol_mismatches: list[dict[str, Any]] = []
    protocol_ids = set(text_protocols) | set(xlsx_protocols)
    protocol_field_totals = []
    for field in PROTOCOL_FIELDS:
        text_total = sum(text_protocols.get(section_id, {}).get(field, 0) for section_id in protocol_ids)
        xlsx_total = sum(xlsx_protocols.get(section_id, {}).get(field, 0) for section_id in protocol_ids)
        protocol_field_totals.append(
            {
                "field": field,
                "text_total": text_total,
                "spreadsheet_total": xlsx_total,
                "diff": text_total - xlsx_total,
            }
        )

    for section_id in sorted(protocol_ids):
        if section_id not in text_protocols:
            protocol_mismatches.append({"section_id": section_id, "field": "", "text_value": "", "spreadsheet_value": "", "diff": "", "issue_type": "missing_from_text_protocols"})
            continue
        if section_id not in xlsx_protocols:
            protocol_mismatches.append({"section_id": section_id, "field": "", "text_value": "", "spreadsheet_value": "", "diff": "", "issue_type": "missing_from_spreadsheet_protocols"})
            continue
        for field in PROTOCOL_FIELDS:
            text_value = text_protocols[section_id][field]
            xlsx_value = xlsx_protocols[section_id][field]
            if text_value != xlsx_value:
                protocol_mismatches.append(
                    {
                        "section_id": section_id,
                        "field": field,
                        "text_value": text_value,
                        "spreadsheet_value": xlsx_value,
                        "diff": text_value - xlsx_value,
                        "issue_type": "protocol_field_mismatch",
                    }
                )

    national_party_rows: dict[int, dict[str, Any]] = defaultdict(lambda: {"party_names": set(), "valid_votes": 0, "paper_votes": 0, "machine_votes": 0})
    for (_region_id, party_id), totals in xlsx_votes.items():
        national_party_rows[party_id]["party_names"].update(totals["party_names"])
        national_party_rows[party_id]["valid_votes"] += totals["valid_votes"]
        national_party_rows[party_id]["paper_votes"] += totals["paper_votes"]
        national_party_rows[party_id]["machine_votes"] += totals["machine_votes"]

    national_spreadsheet_valid_votes = sum(values["valid_votes"] for values in national_party_rows.values())
    national_party_totals = [
        {
            "party_id": party_id,
            "party_names": " | ".join(sorted(totals["party_names"])),
            "spreadsheet_valid_votes": totals["valid_votes"],
            "spreadsheet_paper_votes": totals["paper_votes"],
            "spreadsheet_machine_votes": totals["machine_votes"],
            "spreadsheet_share_of_valid_votes": totals["valid_votes"] / national_spreadsheet_valid_votes if national_spreadsheet_valid_votes else "",
        }
        for party_id, totals in sorted(national_party_rows.items())
    ]

    summary = {
        "spreadsheet_zip": str(spreadsheet_zip),
        "processed_dir": str(processed_dir),
        "workbooks_read": len(workbook_summaries),
        "workbook_summaries": workbook_summaries,
        "section_crosscheck": {
            "text_section_count": len(text_sections),
            "spreadsheet_section_count": len(xlsx_sections),
            "mismatch_count": len(section_mismatches),
        },
        "vote_crosscheck": {
            "region_party_rows": len(vote_crosscheck_rows),
            "mismatch_count": len(vote_mismatches),
            "name_mismatch_count": len(vote_name_mismatches),
            "text_valid_votes": sum(values["valid_votes"] for values in text_votes.values()),
            "spreadsheet_valid_votes": sum(values["valid_votes"] for values in xlsx_votes.values()),
            "diff_valid_votes": sum(values["valid_votes"] for values in text_votes.values()) - sum(values["valid_votes"] for values in xlsx_votes.values()),
            "text_paper_votes": sum(values["paper_votes"] for values in text_votes.values()),
            "spreadsheet_paper_votes": sum(values["paper_votes"] for values in xlsx_votes.values()),
            "diff_paper_votes": sum(values["paper_votes"] for values in text_votes.values()) - sum(values["paper_votes"] for values in xlsx_votes.values()),
            "text_machine_votes": sum(values["machine_votes"] for values in text_votes.values()),
            "spreadsheet_machine_votes": sum(values["machine_votes"] for values in xlsx_votes.values()),
            "diff_machine_votes": sum(values["machine_votes"] for values in text_votes.values()) - sum(values["machine_votes"] for values in xlsx_votes.values()),
        },
        "protocol_crosscheck": {
            "text_protocol_count": len(text_protocols),
            "spreadsheet_protocol_count": len(xlsx_protocols),
            "field_total_rows": len(protocol_field_totals),
            "field_total_mismatches": sum(1 for row in protocol_field_totals if row["diff"] != 0),
            "section_field_mismatches": len(protocol_mismatches),
        },
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(
        output_dir / "spreadsheet_workbook_summaries_2026.csv",
        workbook_summaries,
        ["workbook", "region_id", "section_rows", "vote_party_rows", "protocol_sections"],
    )
    write_csv(
        output_dir / "spreadsheet_vote_crosscheck_2026.csv",
        vote_crosscheck_rows,
        [
            "region_id",
            "party_id",
            "text_party_names",
            "spreadsheet_party_names",
            "text_valid_votes",
            "spreadsheet_valid_votes",
            "diff_valid_votes",
            "text_paper_votes",
            "spreadsheet_paper_votes",
            "diff_paper_votes",
            "text_machine_votes",
            "spreadsheet_machine_votes",
            "diff_machine_votes",
            "text_row_count",
            "spreadsheet_row_count",
        ],
    )
    write_csv(
        output_dir / "spreadsheet_vote_mismatches_2026.csv",
        vote_mismatches,
        [
            "region_id",
            "party_id",
            "text_party_names",
            "spreadsheet_party_names",
            "text_valid_votes",
            "spreadsheet_valid_votes",
            "diff_valid_votes",
            "text_paper_votes",
            "spreadsheet_paper_votes",
            "diff_paper_votes",
            "text_machine_votes",
            "spreadsheet_machine_votes",
            "diff_machine_votes",
            "text_row_count",
            "spreadsheet_row_count",
        ],
    )
    write_csv(
        output_dir / "spreadsheet_vote_name_mismatches_2026.csv",
        vote_name_mismatches,
        ["region_id", "party_id", "text_party_names", "spreadsheet_party_names"],
    )
    write_csv(
        output_dir / "spreadsheet_national_party_totals_2026.csv",
        national_party_totals,
        ["party_id", "party_names", "spreadsheet_valid_votes", "spreadsheet_paper_votes", "spreadsheet_machine_votes", "spreadsheet_share_of_valid_votes"],
    )
    write_csv(
        output_dir / "spreadsheet_section_mismatches_2026.csv",
        section_mismatches,
        ["section_id", "issue_type", "field", "text_value", "spreadsheet_value"],
    )
    write_csv(
        output_dir / "spreadsheet_protocol_field_totals_2026.csv",
        protocol_field_totals,
        ["field", "text_total", "spreadsheet_total", "diff"],
    )
    write_csv(
        output_dir / "spreadsheet_protocol_mismatches_2026.csv",
        protocol_mismatches,
        ["section_id", "issue_type", "field", "text_value", "spreadsheet_value", "diff"],
    )
    (output_dir / "spreadsheet_crosscheck_summary_2026.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--spreadsheet-zip",
        type=Path,
        default=Path("data/raw/cik_2026/spreadsheet.zip"),
        help="Path to CIK spreadsheet.zip.",
    )
    parser.add_argument(
        "--processed-dir",
        type=Path,
        default=Path("data/processed/cik_2026"),
        help="Directory containing normalized CSV outputs.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("outputs/tables"),
        help="Directory for cross-check outputs.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    crosscheck(args.spreadsheet_zip.resolve(), args.processed_dir.resolve(), args.output_dir.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
